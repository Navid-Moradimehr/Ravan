"""Report generation and scheduled exports for industrial data.

Supports:
- CSV, JSON, and Excel export formats
- Scheduled reports (cron-like scheduling)
- Template-based report generation

Open-source alternatives:
- Apache Superset: Full BI platform
- Metabase: Simple BI for databases
- Grafana: Dashboards with reporting
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Callable

logger = logging.getLogger(__name__)

# Try optional dependencies
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import schedule
    SCHEDULE_AVAILABLE = True
except ImportError:
    SCHEDULE_AVAILABLE = False


@dataclass
class ReportTemplate:
    """A report template definition."""
    template_id: str
    name: str
    description: str = ""
    query: str = ""  # SQL query or filter criteria
    format: str = "csv"  # csv, json, xlsx
    schedule: str | None = None  # cron-like: "0 9 * * 1" = weekly Monday 9am
    recipients: list[str] = field(default_factory=list)
    enabled: bool = True
    last_run: str | None = None
    next_run: str | None = None


class ReportEngine:
    """Engine for generating and scheduling reports."""

    def __init__(self, output_dir: str | None = None, template_store_path: str | None = None):
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(__file__), "..", "..", "reports")
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        if template_store_path is None:
            template_store_path = os.getenv("REPORT_TEMPLATE_STORE_PATH")
        self._template_store_path = Path(template_store_path) if template_store_path else None
        self._templates: dict[str, ReportTemplate] = {}
        self._scheduled_jobs: dict[str, Any] = {}
        if self._template_store_path and self._template_store_path.exists():
            self._load_templates()

    def register_template(self, template: ReportTemplate) -> None:
        """Register a report template."""
        self._templates[template.template_id] = template
        logger.info(f"Registered report template: {template.name}")
        self._persist_templates()

    def generate_report(
        self,
        template_id: str,
        start_time: str | None = None,
        end_time: str | None = None,
        format_override: str | None = None,
    ) -> dict[str, Any]:
        """Generate a report from a template."""
        template = self._templates.get(template_id)
        if not template:
            return {"error": f"Template {template_id} not found"}

        # Execute query to get data
        from historian.client import query_sql
        try:
            data = query_sql(template.query, ())
        except Exception as e:
            logger.error(f"Query failed for report {template_id}: {e}")
            return {"error": str(e)}

        # Determine format
        fmt = format_override or template.format

        # Generate filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{template_id}_{timestamp}.{fmt}"
        filepath = self.output_dir / filename

        # Export based on format
        if fmt == "csv":
            self._export_csv(data, filepath)
        elif fmt == "json":
            self._export_json(data, filepath)
        elif fmt == "xlsx":
            if not EXCEL_AVAILABLE:
                return {"error": "openpyxl not installed. Run: pip install openpyxl"}
            self._export_excel(data, filepath)
        else:
            return {"error": f"Unsupported format: {fmt}"}

        # Update template
        template.last_run = datetime.now(timezone.utc).isoformat()

        return {
            "status": "success",
            "template_id": template_id,
            "filename": filename,
            "filepath": str(filepath),
            "format": fmt,
            "record_count": len(data),
            "generated_at": template.last_run,
        }

    def _export_csv(self, data: list[dict], filepath: Path) -> None:
        """Export data to CSV."""
        if not data:
            with open(filepath, "w", newline="") as f:
                pass
            return

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

    def _export_json(self, data: list[dict], filepath: Path) -> None:
        """Export data to JSON."""
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)

    def _export_excel(self, data: list[dict], filepath: Path) -> None:
        """Export data to Excel."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        if data:
            # Headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Data
            for row_idx, row in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header))

        wb.save(filepath)

    def schedule_report(self, template_id: str, cron_expression: str) -> dict[str, Any]:
        """Schedule a report to run periodically."""
        if not SCHEDULE_AVAILABLE:
            return {"error": "schedule library not installed. Run: pip install schedule"}

        template = self._templates.get(template_id)
        if not template:
            return {"error": f"Template {template_id} not found"}

        # Parse simple cron (e.g., "daily at 9:00", "weekly on monday at 9:00")
        # For now, support simple schedule patterns
        import schedule

        job = None
        if cron_expression == "daily":
            job = schedule.every().day.at("09:00")
        elif cron_expression == "hourly":
            job = schedule.every().hour
        elif cron_expression == "weekly":
            job = schedule.every().monday.at("09:00")
        else:
            return {"error": f"Unsupported schedule: {cron_expression}"}

        job.do(self.generate_report, template_id)
        self._scheduled_jobs[template_id] = job

        return {
            "status": "scheduled",
            "template_id": template_id,
            "schedule": cron_expression,
        }

    def list_templates(self) -> list[dict[str, Any]]:
        """List all report templates."""
        return [
            {
                "template_id": t.template_id,
                "name": t.name,
                "description": t.description,
                "format": t.format,
                "schedule": t.schedule,
                "enabled": t.enabled,
                "last_run": t.last_run,
            }
            for t in self._templates.values()
        ]

    def list_generated_reports(self) -> list[dict[str, Any]]:
        """List all generated report files."""
        reports = []
        for file in sorted(self.output_dir.iterdir(), reverse=True):
            if file.is_file():
                stat = file.stat()
                reports.append({
                    "filename": file.name,
                    "filepath": str(file),
                    "size_mb": round(stat.st_size / (1024 * 1024), 2),
                    "created": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
                })
        return reports

    def get_default_templates(self) -> list[ReportTemplate]:
        """Get default report templates."""
        return [
            ReportTemplate(
                template_id="daily_alarms",
                name="Daily Alarms Report",
                description="Summary of all alarms from the last 24 hours",
                query="SELECT * FROM processed_events WHERE severity IN ('warning', 'critical') AND time > NOW() - INTERVAL '24 hours' ORDER BY time DESC",
                format="csv",
                schedule="daily",
            ),
            ReportTemplate(
                template_id="weekly_trends",
                name="Weekly Trends Report",
                description="Trend analysis for the past week",
                query="SELECT asset_id, tag, AVG(value) as avg_value, MIN(value) as min_value, MAX(value) as max_value FROM industrial_events WHERE time > NOW() - INTERVAL '7 days' GROUP BY asset_id, tag",
                format="xlsx",
                schedule="weekly",
            ),
            ReportTemplate(
                template_id="event_summary",
                name="Event Summary",
                description="Summary of all events",
                query="SELECT * FROM industrial_events ORDER BY time DESC LIMIT 1000",
                format="json",
            ),
        ]

    def _load_templates(self) -> None:
        try:
            payload = json.loads(self._template_store_path.read_text(encoding="utf-8"))
        except Exception as exc:  # pragma: no cover - defensive bootstrap path
            raise ValueError(f"failed to load report templates from {self._template_store_path}") from exc

        self._templates = {}
        for item in payload.get("templates", []):
            template = ReportTemplate(
                template_id=item["template_id"],
                name=item["name"],
                description=item.get("description", ""),
                query=item.get("query", ""),
                format=item.get("format", "csv"),
                schedule=item.get("schedule"),
                recipients=list(item.get("recipients", [])),
                enabled=bool(item.get("enabled", True)),
                last_run=item.get("last_run"),
                next_run=item.get("next_run"),
            )
            self._templates[template.template_id] = template

    def _persist_templates(self) -> None:
        if not self._template_store_path:
            return
        self._template_store_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = self._template_store_path.with_suffix(self._template_store_path.suffix + ".tmp")
        payload = {"templates": [asdict(t) for t in self._templates.values()]}
        tmp_path.write_text(json.dumps(payload, indent=2, default=str, sort_keys=True), encoding="utf-8")
        tmp_path.replace(self._template_store_path)


# Global report engine
REPORT_TEMPLATE_STORE_PATH = os.getenv("REPORT_TEMPLATE_STORE_PATH")
report_engine = ReportEngine(template_store_path=REPORT_TEMPLATE_STORE_PATH)

# Load default templates
for template in report_engine.get_default_templates():
    report_engine.register_template(template)
